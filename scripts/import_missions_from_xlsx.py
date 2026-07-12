#!/usr/bin/env python3
from pathlib import Path
import re
import shutil

from openpyxl import load_workbook


def normalize_key(text: str) -> str:
    text = (text or '').strip().lower()
    text = re.sub('[^a-z0-9ãõáéíóúâêîôûàçºªäëïöüñõãÃÕÁÉÍÓÚÂÊÎÔÛÀÇ ]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


project_root = Path(__file__).resolve().parents[1]
plugins_root = Path(__file__).resolve().parents[2]
xlsx_path = project_root / 'missoes_darkorbit_organizado.xlsx'
workspace_output = project_root / 'missions_npc_map_pt.properties'
plugins_output = plugins_root / 'missions_npc_map_pt.properties'

if not xlsx_path.exists():
    print(f"Error: {xlsx_path} not found")
    raise SystemExit(1)

workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
if 'Missões' not in workbook.sheetnames:
    print(f"Error: sheet 'Missões' not found. Available sheets: {workbook.sheetnames}")
    raise SystemExit(1)

sheet = workbook['Missões']
missions = {}

for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    name, desc = row
    if not name or not desc:
        continue
    key = normalize_key(str(name))
    desc_norm = ' | '.join(part.strip() for part in str(desc).split('|') if part.strip())
    if key and key not in missions:
        missions[key] = desc_norm

content_lines = ['# Auto-generated mission -> description mapping (PT from Excel)']
for key in sorted(missions):
    content_lines.append(f'{key}={missions[key]}')

content = '\n'.join(content_lines) + '\n'
workspace_output.write_text(content, encoding='utf-8')
shutil.copyfile(workspace_output, plugins_output)

print(f'Wrote {len(missions)} entries to {workspace_output}')
print(f'Copied to {plugins_output}')
